#!/usr/bin/env python3
"""
Generate formatted AWS billing Excel report matching AWS Bills page structure.
"""

import sys
import json
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


def clean_service_name(service_name):
    """
    Clean AWS service name to match AWS Bills page format.

    Examples:
        "AmazonES" -> "OpenSearch Service"
        "AmazonECS" -> "Elastic Container Service"
        "AmazonQuickSight" -> "QuickSight"
        "AWSLambda" -> "Lambda"

    Args:
        service_name: Original AWS service name from CUR

    Returns:
        Cleaned service name matching AWS Bills page
    """
    # Mapping of CUR service names to AWS Bills display names
    service_name_map = {
        'AmazonES': 'OpenSearch Service',
        'AmazonECS': 'Elastic Container Service',
        'AmazonQuickSight': 'QuickSight',
        'AmazonBedrockService': 'Bedrock Service',
        'AmazonBedrock': 'Bedrock',
        'AmazonRDS': 'Relational Database Service',
        'AmazonEC2': 'Elastic Compute Cloud',
        'AmazonS3': 'Simple Storage Service',
        'AmazonCloudWatch': 'CloudWatch',
        'AmazonDynamoDB': 'DynamoDB',
        'AmazonVPC': 'Virtual Private Cloud',
        'AmazonSNS': 'Simple Notification Service',
        'AmazonSQS': 'Simple Queue Service',
        'AWSLambda': 'Lambda',
        'AWSELB': 'Elastic Load Balancing',
        'AWSCloudTrail': 'CloudTrail',
        'AWSSecretsManager': 'Secrets Manager',
        'AWSCostExplorer': 'Cost Explorer',
        'AmazonRoute53': 'Route 53',
        'AmazonCloudFront': 'CloudFront',
        'AmazonApiGateway': 'API Gateway',
        'AmazonECR': 'EC2 Container Registry (ECR)',
        'AmazonEKS': 'Elastic Container Service for Kubernetes',
        'AWSGlue': 'Glue',
        'AWSQueueService': 'Simple Queue Service',
        'AmazonCognito': 'Cognito',
        'AmazonECRPublic': 'EC2 Container Registry (Public)',
        'AmazonGuardDuty': 'GuardDuty',
        'AmazonNeptune': 'Neptune',
        'awskms': 'Key Management Service',
        'AWSKMS': 'Key Management Service',
        'DataTransfer': 'Data Transfer',
        'Kiro': 'Kiro',
    }

    # Return mapped name if exists, otherwise clean the original name
    if service_name in service_name_map:
        return service_name_map[service_name]

    # Fallback: Remove "Amazon " and "AWS " prefixes
    cleaned = re.sub(r'^Amazon\s*', '', service_name)
    cleaned = re.sub(r'^AWS\s*', '', cleaned)
    return cleaned


def apply_professional_style(ws, header_row=1, data_start_row=2):
    """
    Apply Microsoft 365 professional styling with Aptos font.

    Args:
        ws: Worksheet object
        header_row: Row number for headers (1-indexed)
        data_start_row: First row of data (1-indexed)
    """
    # Define styles
    header_font = Font(name='Aptos', size=11, bold=True)
    body_font = Font(name='Aptos', size=11)
    header_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    body_alignment = Alignment(horizontal='right', vertical='center')
    text_alignment = Alignment(horizontal='left', vertical='center')

    # Apply header styles
    for cell in ws[header_row]:
        if cell.value:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

    # Apply body styles
    max_row = ws.max_row
    max_col = ws.max_column

    for row in range(data_start_row, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = body_font

            # First column (Description) is left-aligned, others are right-aligned
            if col == 1:
                cell.alignment = text_alignment
            else:
                cell.alignment = body_alignment
                # Format numbers as currency
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'

    # Auto-adjust column widths
    for col in range(1, max_col + 1):
        column_letter = get_column_letter(col)
        max_length = 0
        for row in range(1, max_row + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value:
                cell_length = len(str(cell.value))
                max_length = max(max_length, cell_length)
        # Set width with padding
        ws.column_dimensions[column_letter].width = min(max_length + 3, 20)

    # Make Description column wider
    ws.column_dimensions['A'].width = 35


def create_service_summary_sheet(wb, billing_data_list):
    """
    Create Sheet1: Service-level summary matching AWS Bills page.
    Each row is a service, each column is a month.

    Args:
        wb: Workbook object
        billing_data_list: List of billing data dicts, one per year
    """
    ws = wb.create_sheet("Service Summary")

    # Month columns
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

    # Write header row
    ws.append(["Description"] + months + ["Total"])

    # Collect all unique services across all years
    all_services = set()
    for billing_data in billing_data_list:
        all_services.update(billing_data["by_service"].keys())

    # Sort services alphabetically by cleaned name
    sorted_services = sorted(all_services, key=lambda s: clean_service_name(s))

    # Write data for each year
    for billing_data in billing_data_list:
        year = billing_data["metadata"]["year"]
        currency = billing_data["metadata"]["currency"]

        # Add year header if multiple years
        if len(billing_data_list) > 1:
            ws.append([f"--- {year} ---"])

        # Add each service with costs
        for service in sorted_services:
            service_data = billing_data["by_service"].get(service, {month: 0 for month in months})
            monthly_costs = [service_data.get(month, 0) for month in months]

            # Only include services with non-zero costs
            if sum(monthly_costs) > 0:
                row_num = ws.max_row + 1
                cleaned_name = clean_service_name(service)
                ws.append([cleaned_name] + monthly_costs + [f"=SUM(B{row_num}:M{row_num})"])

        # Add year total
        if len(billing_data_list) > 1:
            # Find the start row for this year's data
            year_start_row = ws.max_row - sum(1 for s in sorted_services if sum(billing_data["by_service"].get(s, {month: 0 for month in months}).values()) > 0) + 1
        else:
            year_start_row = 2

        year_end_row = ws.max_row
        total_row_num = ws.max_row + 1

        total_label = f"Total" if len(billing_data_list) == 1 else f"{year} Total"
        total_row = [total_label]
        for col_letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
            total_row.append(f"=SUM({col_letter}{year_start_row}:{col_letter}{year_end_row})")

        ws.append(total_row)

        # Add empty row between years if multiple years
        if len(billing_data_list) > 1:
            ws.append([])

    # Apply professional styling
    apply_professional_style(ws)


def extract_region_from_usage_type(usage_type, description, region_code=''):
    """
    Extract region name from usage type, description, or region code.

    Args:
        usage_type: Usage type string (e.g., "USW2-DataTransfer-Regional-Bytes")
        description: Description string
        region_code: AWS region code (e.g., "us-west-2")

    Returns:
        Region name (e.g., "US West (Oregon)")
    """
    # Region code mapping (both CUR codes and AWS region codes)
    region_map = {
        'USE1': 'US East (N. Virginia)',
        'us-east-1': 'US East (N. Virginia)',
        'USE2': 'US East (Ohio)',
        'us-east-2': 'US East (Ohio)',
        'USW1': 'US West (N. California)',
        'us-west-1': 'US West (N. California)',
        'USW2': 'US West (Oregon)',
        'us-west-2': 'US West (Oregon)',
        'APN1': 'Asia Pacific (Tokyo)',
        'ap-northeast-1': 'Asia Pacific (Tokyo)',
        'APN2': 'Asia Pacific (Seoul)',
        'ap-northeast-2': 'Asia Pacific (Seoul)',
        'APN3': 'Asia Pacific (Osaka)',
        'ap-northeast-3': 'Asia Pacific (Osaka)',
        'APS1': 'Asia Pacific (Singapore)',
        'ap-southeast-1': 'Asia Pacific (Singapore)',
        'APS2': 'Asia Pacific (Sydney)',
        'ap-southeast-2': 'Asia Pacific (Sydney)',
        'APS3': 'Asia Pacific (Mumbai)',
        'ap-south-1': 'Asia Pacific (Mumbai)',
        'CAN1': 'Canada (Central)',
        'ca-central-1': 'Canada (Central)',
        'EU': 'EU (Ireland)',
        'eu-west-1': 'EU (Ireland)',
        'EUC1': 'EU (Frankfurt)',
        'eu-central-1': 'EU (Frankfurt)',
        'EUN1': 'EU (Stockholm)',
        'eu-north-1': 'EU (Stockholm)',
        'EUW2': 'EU (London)',
        'eu-west-2': 'EU (London)',
        'EUW3': 'EU (Paris)',
        'eu-west-3': 'EU (Paris)',
        'SAE1': 'South America (Sao Paulo)',
        'sa-east-1': 'South America (Sao Paulo)',
    }

    # Try region_code first
    if region_code and region_code in region_map:
        return region_map[region_code]

    # Try to extract region code from usage type
    for code, name in region_map.items():
        if usage_type.startswith(code):
            return name

    # Try to find region in description
    for code, name in region_map.items():
        if name in description:
            return name

    return 'Global'


def create_usage_details_sheet(wb, billing_data_list, target_month='Feb'):
    """
    Create Sheet2: Detailed usage breakdown for drill-down.
    Shows usage types grouped by service and region.

    Args:
        wb: Workbook object
        billing_data_list: List of billing data dicts, one per year
        target_month: Month to show details for (default: Feb)
    """
    ws = wb.create_sheet("Usage Details")

    # Write header row
    ws.append(["Service", "Region", "Usage Type", "Description", "Quantity", "Rate", "Cost"])

    # Get data for the target year (most recent)
    billing_data = billing_data_list[-1]

    if 'usage_details' not in billing_data or not billing_data['usage_details']:
        # No detail data available
        ws.append(["No detailed usage data available"])
        apply_professional_style(ws)
        return

    # Filter for target month and group by service
    usage_details = [d for d in billing_data['usage_details'] if d['month'] == target_month]

    # Group by service
    from collections import defaultdict
    by_service = defaultdict(list)
    for detail in usage_details:
        service = detail['service']
        by_service[service].append(detail)

    # Sort services by total cost
    service_totals = {service: sum(d['cost'] for d in details)
                     for service, details in by_service.items()}
    sorted_services = sorted(service_totals.keys(), key=lambda s: service_totals[s], reverse=True)

    # Write data for each service
    for service in sorted_services:
        details = by_service[service]
        service_total = service_totals[service]

        # Skip services with zero cost
        if service_total < 0.01:
            continue

        # Add service header
        cleaned_service_name = clean_service_name(service)
        ws.append([f"--- {cleaned_service_name} (${service_total:.2f}) ---"])

        # Group by region within service
        by_region = defaultdict(list)
        for detail in details:
            region_code = detail.get('region', '')
            # Convert region code to friendly name
            region = extract_region_from_usage_type(
                detail.get('usage_type', ''),
                detail.get('description', ''),
                region_code
            )
            by_region[region].append(detail)

        # Calculate region totals
        region_totals = {region: sum(d['cost'] for d in region_details)
                        for region, region_details in by_region.items()}

        # Sort regions by cost
        sorted_regions = sorted(region_totals.keys(), key=lambda r: region_totals[r], reverse=True)

        # Write data for each region
        for region in sorted_regions:
            region_details = by_region[region]
            region_total = region_totals[region]

            # Add region header if region exists and there are multiple regions
            if region and region != 'Global':
                ws.append(["", region, "", "", "", "", region_total])

            # Sort usage types within region by cost
            region_details.sort(key=lambda d: d['cost'], reverse=True)

            # Add each usage type
            for detail in region_details:
                usage_type = detail.get('usage_type', '')
                description = detail.get('description', '')
                quantity = detail.get('quantity', 0)
                rate = detail.get('rate', 0)
                cost = detail.get('cost', 0)

                # Format quantity with unit (assume GB for data transfer, otherwise use generic unit)
                if quantity > 0:
                    if 'Bytes' in usage_type or 'GB' in description:
                        quantity_str = f"{quantity:.3f} GB"
                    else:
                        quantity_str = f"{quantity:.3f}"
                else:
                    quantity_str = "0 GB" if 'Bytes' in usage_type else "0"

                # Format rate
                if rate > 0:
                    if 'Bytes' in usage_type or 'GB' in description:
                        rate_str = f"${rate:.4f} per GB"
                    else:
                        rate_str = f"${rate:.4f}"
                else:
                    rate_str = "$0.00"

                # Add service prefix to usage type for better readability (matching AWS Bills format)
                display_usage_type = usage_type
                if service == 'DataTransfer' or 'DataTransfer' in usage_type:
                    if not usage_type.startswith('AWS Data Transfer'):
                        display_usage_type = f"AWS Data Transfer {usage_type}"
                elif service == 'AmazonEC2' and 'BoxUsage' in usage_type:
                    display_usage_type = f"Amazon EC2 {usage_type}"
                elif service == 'AmazonS3':
                    display_usage_type = f"Amazon S3 {usage_type}"

                ws.append([
                    "",  # Service (blank, already in header)
                    "",  # Region (blank unless it's a region header)
                    display_usage_type,
                    description,
                    quantity_str,
                    rate_str,
                    cost
                ])

        # Add blank row between services
        ws.append([])

    # Apply professional styling
    apply_professional_style(ws)

    # Adjust column widths
    ws.column_dimensions['A'].width = 30  # Service
    ws.column_dimensions['B'].width = 30  # Region
    ws.column_dimensions['C'].width = 45  # Usage Type
    ws.column_dimensions['D'].width = 80  # Description
    ws.column_dimensions['E'].width = 15  # Quantity
    ws.column_dimensions['F'].width = 12  # Rate
    ws.column_dimensions['G'].width = 12  # Cost


def generate_report(billing_data_list, output_file):
    """
    Generate Excel report with service summary and usage details.

    Args:
        billing_data_list: List of billing data dicts, one per year
        output_file: Output Excel file path
    """
    # Create workbook
    wb = Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Create Sheet1: Service Summary
    create_service_summary_sheet(wb, billing_data_list)

    # Create Sheet2: Usage Details (drill-down)
    create_usage_details_sheet(wb, billing_data_list)

    # Save workbook
    wb.save(output_file)
    print(f"Report saved to: {output_file}")


def main():
    """Main entry point for the script."""
    if len(sys.argv) < 3:
        print("Usage: python generate_excel_report.py <output_file.xlsx> <billing_data1.json> [<billing_data2.json> ...]", file=sys.stderr)
        print("Example: python generate_excel_report.py report.xlsx billing_2026.json", file=sys.stderr)
        sys.exit(1)

    output_file = sys.argv[1]
    billing_files = sys.argv[2:]

    # Load billing data from JSON files
    billing_data_list = []
    for file_path in billing_files:
        try:
            with open(file_path, 'r') as f:
                data = json.load(f)
                billing_data_list.append(data)
        except FileNotFoundError:
            print(f"Error: File not found: {file_path}", file=sys.stderr)
            sys.exit(1)
        except json.JSONDecodeError:
            print(f"Error: Invalid JSON in file: {file_path}", file=sys.stderr)
            sys.exit(1)

    # Sort by year
    billing_data_list.sort(key=lambda x: x["metadata"]["year"])

    # Generate report
    generate_report(billing_data_list, output_file)

    print(f"\nSuccessfully generated report with {len(billing_data_list)} year(s) of data")


if __name__ == "__main__":
    main()
